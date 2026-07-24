#!/usr/bin/env python3
"""Compare stat-distribution models on a deliberately narrow item cohort.

This is a diagnostic, not a claim that Blizzard's listed ItemLevel is perfect.
The default cohort uses pure-stat chest armor with unit-cost stats so slot and
stat coefficients do not confound the exponent comparison.
"""

from __future__ import annotations

import argparse
import math
import statistics
from pathlib import Path

from openpyxl import load_workbook


CURRENT_EXPONENT = math.log(2) / math.log(1.5)
LOG2_THREE = math.log(3) / math.log(2)
UNIT_COST_STAT_IDS = {3, 4, 5, 6, 12, 13, 14, 15, 21, 31, 32, 35, 36, 37, 44}


def quality_mod(quality: int, level: int) -> float:
    if quality == 4:
        if level >= 200:
            return 1.320 * level - 120
        if level >= 100:
            return 0.700 * level - 2
        return 0.689 * level + 1
    if quality == 3:
        if level >= 136:
            return 0.880 * level - 39.25
        if level >= 80:
            return 0.674 * level - 8
        return 0.641 * level - 4
    if quality == 2:
        if level >= 130:
            return 0.801 * level - 38.3
        if level >= 80:
            return 0.505 * level - 4.5
        return 0.495 * level - 2.85
    raise ValueError(f"Unsupported quality: {quality}")


def first_level_meeting(quality: int, required_quality_budget: float) -> int | None:
    for level in range(1, 301):
        value = quality_mod(quality, level)
        if value > 0 and value >= required_quality_budget:
            return level
    return None


def lp_required_budget(values: list[float], exponent: float) -> float:
    return sum(value**exponent for value in values) ** (1 / exponent)


def distribution_factor(values: list[float]) -> float:
    total = sum(values)
    shares = [value / total for value in values]
    return sum(share**1.5 for share in shares) ** (2 / 3)


def predictions(values: list[float], quality: int) -> dict[str, int | None]:
    total = sum(values)
    factor = distribution_factor(values)
    required = {
        "current p=1.709511": lp_required_budget(values, CURRENT_EXPONENT),
        "alternative p=1.5": lp_required_budget(values, 1.5),
        "misquoted p=log2(3)": lp_required_budget(values, LOG2_THREE),
        "peer formula as written": total / factor,
        "reciprocal-corrected 3/2 norm": total * factor,
    }
    return {
        name: first_level_meeting(quality, budget)
        for name, budget in required.items()
    }


def load_cohort(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["all items"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {name: position for position, name in enumerate(headers)}
    cohort: list[dict[str, object]] = []

    for row in rows:
        quality = row[index["Quality"]]
        level = row[index["ItemLevel"]]
        if row[index["class"]] != 4 or row[index["InventoryType"]] != 5:
            continue
        if quality not in (2, 3, 4) or not isinstance(level, (int, float)):
            continue
        if not 1 <= level <= 300:
            continue
        if any(row[index[f"spellid_{number}"]] not in (None, 0) for number in range(1, 6)):
            continue
        if any(row[index[f"socketColor_{number}"]] not in (None, 0) for number in range(1, 4)):
            continue
        if any(row[index[name]] not in (None, 0) for name in (
            "holy_res", "fire_res", "nature_res", "frost_res", "shadow_res", "arcane_res"
        )):
            continue

        values: list[float] = []
        valid = True
        for number in range(1, 11):
            stat_type = row[index[f"stat_type{number}"]]
            stat_value = row[index[f"stat_value{number}"]]
            if stat_type in (None, 0) and stat_value in (None, 0):
                continue
            if stat_type not in UNIT_COST_STAT_IDS:
                valid = False
                break
            if not isinstance(stat_value, (int, float)) or stat_value <= 0:
                valid = False
                break
            values.append(float(stat_value))

        if not valid or not values:
            continue

        cohort.append({
            "entry": row[index["entry"]],
            "actual": int(level),
            "quality": int(quality),
            "values": values,
        })

    workbook.close()
    return cohort


def summarize(errors: list[int]) -> str:
    absolute = [abs(error) for error in errors]
    rmse = math.sqrt(sum(error * error for error in errors) / len(errors))
    within_two = sum(value <= 2 for value in absolute) / len(absolute)
    return (
        f"n={len(errors):3d}  MAE={statistics.fmean(absolute):6.3f}  "
        f"RMSE={rmse:6.3f}  bias={statistics.fmean(errors):+6.3f}  "
        f"within ±2={within_two:6.1%}"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=Path("Data/item_template_pruned.xlsm"),
    )
    args = parser.parse_args()

    cohort = load_cohort(args.workbook)
    if not cohort:
        raise SystemExit("No rows matched the benchmark cohort.")

    errors: dict[str, list[int]] = {}
    failures: dict[str, int] = {}
    for item in cohort:
        item_predictions = predictions(item["values"], item["quality"])
        for name, predicted in item_predictions.items():
            if predicted is None:
                failures[name] = failures.get(name, 0) + 1
                continue
            errors.setdefault(name, []).append(predicted - item["actual"])

    print("Cohort: pure-stat, socket-free, proc-free chest armor using unit-cost stats")
    print(f"Rows: {len(cohort)}")
    print("Listed ItemLevel is treated as a noisy diagnostic label, not ground truth.\n")
    for name, model_errors in errors.items():
        suffix = f"  failures={failures.get(name, 0)}" if failures.get(name, 0) else ""
        print(f"{name:34s} {summarize(model_errors)}{suffix}")


if __name__ == "__main__":
    main()
